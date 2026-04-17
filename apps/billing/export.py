"""
Export utilities for generating PDF and Excel reports.

Provides functions to export invoices, reports, and other data
to PDF and Excel formats.
"""
from io import BytesIO, StringIO
from datetime import datetime
import logging

from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

from apps.billing.models import Invoice
from apps.reports.models import SalesReport, ProfitAndLoss

logger = logging.getLogger(__name__)


class PDFExporter:
    """Generate PDF reports and invoices."""
    
    @staticmethod
    def export_invoice(invoice):
        """
        Export an invoice to PDF.
        
        Args:
            invoice: Invoice instance to export
            
        Returns:
            HttpResponse with PDF content
        """
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#007bff'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Add title
            elements.append(Paragraph(f'Invoice #{invoice.invoice_number}', title_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Add invoice details
            details = f"""
            <b>Invoice Date:</b> {invoice.invoice_date}<br/>
            <b>Due Date:</b> {invoice.due_date}<br/>
            <b>Status:</b> {invoice.get_status_display()}<br/>
            <br/>
            <b>Bill To:</b><br/>
            {invoice.customer.name}<br/>
            {invoice.customer.address}<br/>
            {invoice.customer.city}, {invoice.customer.state} {invoice.customer.postal_code}<br/>
            """
            elements.append(Paragraph(details, styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Add line items table
            data = [['Item', 'Qty', 'Unit Price', 'Amount']]
            for item in invoice.items.all():
                data.append([
                    item.product.name,
                    str(item.quantity),
                    f"${item.unit_price:.2f}",
                    f"${item.line_total:.2f}"
                ])
            
            # Add totals
            data.append(['', '', 'Subtotal:', f"${invoice.subtotal:.2f}"])
            data.append(['', '', 'Tax:', f"${invoice.tax_amount:.2f}"])
            data.append(['', '', 'Total:', f"${invoice.total_amount:.2f}"])
            data.append(['', '', 'Paid:', f"${invoice.paid_amount:.2f}"])
            data.append(['', '', 'Balance:', f"${invoice.remaining_balance:.2f}"])
            
            table = Table(data, colWidths=[3*inch, 0.8*inch, 1.2*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -5), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, -5), (-1, -1), 'Helvetica-Bold'),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Add notes
            if invoice.notes:
                elements.append(Paragraph(f'<b>Notes:</b> {invoice.notes}', styles['Normal']))
            
            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.invoice_number}.pdf"'
            return response
        except Exception as e:
            logger.error(f"Error exporting invoice to PDF: {str(e)}")
            raise
    
    @staticmethod
    def export_sales_report(start_date, end_date):
        """
        Export sales report to PDF.
        
        Args:
            start_date: Start date for report
            end_date: End date for report
            
        Returns:
            HttpResponse with PDF content
        """
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#007bff'),
                spaceAfter=30,
                alignment=1
            )
            elements.append(Paragraph('Sales Report', title_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Date range
            date_range = f"Period: {start_date} to {end_date}"
            elements.append(Paragraph(date_range, styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Get data and build table
            reports = SalesReport.objects.filter(
                report_date__gte=start_date,
                report_date__lte=end_date
            ).order_by('report_date')
            
            data = [['Date', 'Sales', 'Orders', 'Avg Order Value', 'Tax']]
            for report in reports:
                data.append([
                    str(report.report_date),
                    f"${report.total_sales:.2f}",
                    str(report.total_orders),
                    f"${report.average_order_value:.2f}",
                    f"${report.tax_collected:.2f}"
                ])
            
            table = Table(data, colWidths=[1.2*inch, 1.2*inch, 1*inch, 1.5*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_to_{end_date}.pdf"'
            return response
        except Exception as e:
            logger.error(f"Error exporting sales report to PDF: {str(e)}")
            raise


class ExcelExporter:
    """Generate Excel reports and exports."""
    
    @staticmethod
    def export_invoices(invoice_ids):
        """
        Export multiple invoices to Excel.
        
        Args:
            invoice_ids: List of invoice IDs to export
            
        Returns:
            HttpResponse with Excel content
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Invoices'
            
            # Define styles
            header_fill = PatternFill(start_color='007bff', end_color='007bff', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add headers
            headers = ['Invoice #', 'Customer', 'Date', 'Due Date', 'Amount', 'Paid', 'Balance', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            invoices = Invoice.objects.filter(id__in=invoice_ids)
            for row, invoice in enumerate(invoices, 2):
                ws.cell(row=row, column=1).value = invoice.invoice_number
                ws.cell(row=row, column=2).value = invoice.customer.name
                ws.cell(row=row, column=3).value = invoice.invoice_date
                ws.cell(row=row, column=4).value = invoice.due_date
                ws.cell(row=row, column=5).value = float(invoice.total_amount)
                ws.cell(row=row, column=6).value = float(invoice.paid_amount)
                ws.cell(row=row, column=7).value = float(invoice.remaining_balance)
                ws.cell(row=row, column=8).value = invoice.get_status_display()
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            
            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="invoices_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            return response
        except Exception as e:
            logger.error(f"Error exporting invoices to Excel: {str(e)}")
            raise
    
    @staticmethod
    def export_sales_data(start_date, end_date):
        """
        Export sales data to Excel with statistics.
        
        Args:
            start_date: Start date for export
            end_date: End date for export
            
        Returns:
            HttpResponse with Excel content
        """
        try:
            # Create workbook with multiple sheets
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Sheet 1: Daily Summary
            ws_summary = wb.create_sheet('Daily Summary')
            reports = SalesReport.objects.filter(
                report_date__gte=start_date,
                report_date__lte=end_date
            ).order_by('report_date')
            
            # Headers
            headers = ['Date', 'Total Sales', 'Orders', 'Avg Order', 'Tax']
            for col, header in enumerate(headers, 1):
                ws_summary.cell(row=1, column=col).value = header
            
            # Data
            for row, report in enumerate(reports, 2):
                ws_summary.cell(row=row, column=1).value = report.report_date
                ws_summary.cell(row=row, column=2).value = float(report.total_sales)
                ws_summary.cell(row=row, column=3).value = report.total_orders
                ws_summary.cell(row=row, column=4).value = float(report.average_order_value)
                ws_summary.cell(row=row, column=5).value = float(report.tax_collected)
            
            # Adjust column widths
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws_summary.column_dimensions[col].width = 15
            
            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="sales_data_{start_date}_to_{end_date}.xlsx"'
            return response
        except Exception as e:
            logger.error(f"Error exporting sales data to Excel: {str(e)}")
            raise
